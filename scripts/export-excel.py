#!/usr/bin/env python3
"""
export-excel.py — Ekspor bills.csv ke Excel, dipisah per bulan.

Menghasilkan satu file .xlsx berisi:
  - Sheet "Ringkasan": pemasukan / pengeluaran / saldo per bulan + Grand Total
  - Satu sheet per bulan (nama "YYYY-MM"), berisi transaksi bulan itu saja,
    dengan baris Total Pemasukan, Total Pengeluaran, dan Saldo.

Usage:
  python3 scripts/export-excel.py                 # simpan ke data/backups/bills_export_YYYYMMDD.xlsx
  python3 scripts/export-excel.py --out FILE.xlsx # simpan ke path lain
"""

import argparse
import csv
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
BILLS_CSV = DATA_DIR / "bills.csv"
BACKUP_DIR = DATA_DIR / "backups"

RETENSI_HARI = 30  # hapus file bills_export_*.xlsx lebih tua dari ini tiap export baru dibuat

# Kolom yang ditulis ke tiap sheet bulan (urutan tetap).
# `no_resi` sengaja ditaruh paling kiri agar mudah dilihat; `waktu` setelah tanggal.
# Urutan di Excel boleh beda dari urutan di CSV (penulisan sheet berdasarkan nama kolom).
COLUMNS = [
    "no_resi",
    "tanggal",
    "waktu",
    "tipe",
    "kategori",
    "tempat",
    "item",
    "jumlah",
    "catatan",
    "channel",
    "pencatat",
]

NO_DATE_SHEET = "Tanpa Tanggal"

# ── Styling ──────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
TOTAL_FONT = Font(bold=True)
POS_FONT = Font(bold=True, color="1B7F37")   # hijau: saldo positif
NEG_FONT = Font(bold=True, color="C0392B")   # merah: saldo negatif
RUPIAH_FMT = "#,##0"


def cleanup_old_exports(backup_dir: Path, days: int = RETENSI_HARI) -> int:
    """Hapus file bills_export_*.xlsx yang lebih tua dari `days` hari.

    Dipanggil setiap kali export baru berhasil dibuat, terlepas apakah dipicu
    langsung (`python3 scripts/export-excel.py`) atau lewat `backup.sh` —
    retensi tidak bergantung pada entry point mana yang dipakai.
    Return jumlah file yang dihapus.
    """
    if not backup_dir.exists():
        return 0
    cutoff = datetime.now() - timedelta(days=days)
    deleted = 0
    for f in backup_dir.glob("bills_export_*.xlsx"):
        try:
            if datetime.fromtimestamp(f.stat().st_mtime) < cutoff:
                f.unlink()
                deleted += 1
        except Exception:
            pass  # jangan sampai kegagalan cleanup menggagalkan export itu sendiri
    return deleted


def load_bills() -> list[dict]:
    """Load data dari bills.csv (kosong jika tidak ada)."""
    if not BILLS_CSV.exists():
        return []
    try:
        with open(BILLS_CSV, encoding="utf-8", newline="") as f:
            return list(csv.DictReader(f))
    except Exception:
        return []


def row_tipe(row: dict) -> str:
    """Ambil tipe transaksi, tahan-mundur ke 'pengeluaran' untuk data lama."""
    return (row.get("tipe") or "pengeluaran").strip().lower()


def row_jumlah(row: dict) -> float:
    try:
        return float(row.get("jumlah") or 0)
    except (ValueError, TypeError):
        return 0.0


def month_key(row: dict) -> str:
    """Kunci bulan 'YYYY-MM' dari kolom tanggal, atau NO_DATE_SHEET jika invalid."""
    tanggal = (row.get("tanggal") or "").strip()
    try:
        return datetime.strptime(tanggal, "%Y-%m-%d").strftime("%Y-%m")
    except ValueError:
        # Coba ambil prefix YYYY-MM kalau formatnya masih masuk akal
        if len(tanggal) >= 7 and tanggal[:4].isdigit() and tanggal[5:7].isdigit():
            return tanggal[:7]
        return NO_DATE_SHEET


def format_rupiah(amount: float) -> str:
    return f"Rp {amount:,.0f}".replace(",", ".")


def summary_sort_key(name: str) -> tuple[int, str]:
    """Urutan bulan: kronologis, dengan NO_DATE_SHEET selalu paling akhir."""
    return (1, name) if name == NO_DATE_SHEET else (0, name)


def autofit(ws) -> None:
    """Set lebar kolom sederhana berdasarkan panjang teks maksimum."""
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if cell.number_format == RUPIAH_FMT and isinstance(cell.value, (int, float)):
                length = len(format_rupiah(cell.value))
            widths[cell.column] = max(widths.get(cell.column, 0), length)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 40)


def write_month_sheet(
    wb: Workbook, sheet_name: str, rows: list[dict], saldo_masuk: float | None = None
) -> tuple[float, float]:
    """Tulis satu sheet bulan. Return (total_pemasukan, total_pengeluaran).

    `saldo_masuk` = saldo kas di akhir bulan sebelumnya. None untuk sheet yang tidak
    punya posisi kronologis, sehingga baris Saldo Awal/Akhir Bulan dilewati.
    """
    ws = wb.create_sheet(title=sheet_name)

    # Header
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    jumlah_col = COLUMNS.index("jumlah") + 1  # 1-indexed untuk openpyxl

    # Urutkan: tanggal naik, lalu pemasukan sebelum pengeluaran
    ordered = sorted(
        rows,
        key=lambda r: ((r.get("tanggal") or ""), 0 if row_tipe(r) == "pemasukan" else 1),
    )

    total_in = 0.0
    total_out = 0.0
    for row in ordered:
        ws.append([row.get(col, "") for col in COLUMNS])
        cell = ws.cell(row=ws.max_row, column=jumlah_col)
        cell.value = row_jumlah(row)
        cell.number_format = RUPIAH_FMT
        if row_tipe(row) == "pemasukan":
            total_in += row_jumlah(row)
        else:
            total_out += row_jumlah(row)

    saldo = total_in - total_out
    label_col = jumlah_col - 1 if jumlah_col > 1 else 1

    # Baris pemisah (list penuh berisi None supaya benar-benar menambah baris baru)
    ws.append([None] * len(COLUMNS))

    def total_row(label: str, value: float, font: Font) -> None:
        rowvals: list = [None] * len(COLUMNS)
        rowvals[label_col - 1] = label
        rowvals[jumlah_col - 1] = value
        ws.append(rowvals)
        r = ws.max_row
        ws.cell(row=r, column=label_col).font = font
        vc = ws.cell(row=r, column=jumlah_col)
        vc.number_format = RUPIAH_FMT
        vc.font = font

    total_row("Total Pemasukan", total_in, TOTAL_FONT)
    total_row("Total Pengeluaran", total_out, TOTAL_FONT)
    total_row("Arus Kas Bulan", saldo, POS_FONT if saldo >= 0 else NEG_FONT)

    # Saldo = sisa kas kumulatif, dibawa dari bulan sebelumnya. Sheet tanpa posisi
    # kronologis (NO_DATE_SHEET) tidak ikut rantai ini, ditandai saldo_masuk=None.
    if saldo_masuk is not None:
        saldo_keluar = saldo_masuk + saldo
        total_row("Saldo Awal Bulan", saldo_masuk, POS_FONT if saldo_masuk >= 0 else NEG_FONT)
        total_row("Saldo Akhir Bulan", saldo_keluar, POS_FONT if saldo_keluar >= 0 else NEG_FONT)

    autofit(ws)
    return total_in, total_out


def write_summary_sheet(wb: Workbook, per_month: dict[str, tuple[float, float, int]]) -> None:
    """Sheet Ringkasan: arus kas per bulan + saldo kas kumulatif + Grand Total.

    `Arus Kas Bulan` berdiri sendiri per bulan; `Saldo Akhir` adalah sisa kas kumulatif
    yang dibawa antar bulan — dua angka berbeda yang dulu sama-sama dinamai "Saldo".
    """
    ws = wb.create_sheet(title="Ringkasan", index=0)
    headers = [
        "Bulan", "Jumlah Transaksi", "Total Pemasukan", "Total Pengeluaran",
        "Arus Kas Bulan", "Saldo Akhir",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    grand_in = grand_out = grand_count = 0
    saldo = 0.0
    saldo_terakhir = 0.0
    for bulan in sorted(per_month, key=summary_sort_key):
        total_in, total_out, count = per_month[bulan]
        arus = total_in - total_out
        # NO_DATE_SHEET tidak punya posisi kronologis → tidak ikut rantai saldo kumulatif
        kronologis = bulan != NO_DATE_SHEET
        if kronologis:
            saldo += arus
            saldo_terakhir = saldo
        ws.append([bulan, count, total_in, total_out, arus, saldo if kronologis else None])
        r = ws.max_row
        ws.cell(row=r, column=3).number_format = RUPIAH_FMT
        ws.cell(row=r, column=4).number_format = RUPIAH_FMT
        ac = ws.cell(row=r, column=5)
        ac.number_format = RUPIAH_FMT
        ac.font = POS_FONT if arus >= 0 else NEG_FONT
        if kronologis:
            kc = ws.cell(row=r, column=6)
            kc.number_format = RUPIAH_FMT
            kc.font = POS_FONT if saldo >= 0 else NEG_FONT
        grand_in += total_in
        grand_out += total_out
        grand_count += count

    # Grand Total. Kolom Saldo Akhir memakai saldo bulan terakhir — menjumlahkan kolom
    # kumulatif tidak bermakna.
    grand_arus = grand_in - grand_out
    ws.append(["GRAND TOTAL", grand_count, grand_in, grand_out, grand_arus, saldo_terakhir])
    r = ws.max_row
    for col in range(1, 7):
        ws.cell(row=r, column=col).font = TOTAL_FONT
    ws.cell(row=r, column=3).number_format = RUPIAH_FMT
    ws.cell(row=r, column=4).number_format = RUPIAH_FMT
    ga = ws.cell(row=r, column=5)
    ga.number_format = RUPIAH_FMT
    ga.font = POS_FONT if grand_arus >= 0 else NEG_FONT
    gs = ws.cell(row=r, column=6)
    gs.number_format = RUPIAH_FMT
    gs.font = POS_FONT if saldo_terakhir >= 0 else NEG_FONT

    autofit(ws)


# Kolom penentu "duplikat": semua kolom data KECUALI no_resi.
# Dengan begitu, item-item berbeda dalam satu struk (yang berbagi no_resi sama) TIDAK
# salah-ditandai duplikat, tapi transaksi yang tercatat dua kali tetap tertangkap
# walau no_resi-nya berbeda.
DUP_KEY_COLUMNS = [c for c in COLUMNS if c != "no_resi"]


def find_duplicates(bills: list[dict]) -> list[list[dict]]:
    """Cari grup baris yang isinya identik (semua kolom sama kecuali no_resi).

    Return daftar grup (tiap grup = list baris, >1 baris). Bersifat laporan saja —
    tidak ada baris yang dibuang. Urutan grup mengikuti kemunculan pertama.
    """
    groups: dict[tuple, list[dict]] = {}
    order: list[tuple] = []
    for row in bills:
        key = tuple(norm_dup_value(col, row.get(col)) for col in DUP_KEY_COLUMNS)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(row)
    return [groups[k] for k in order if len(groups[k]) > 1]


def norm_dup_value(col: str, value) -> str:
    """Normalisasi nilai untuk pembandingan duplikat (jumlah dibandingkan sbagai angka)."""
    if col == "jumlah":
        return str(row_jumlah({"jumlah": value}))
    return str(value or "").strip()


def write_duplicate_sheet(wb: Workbook, groups: list[list[dict]]) -> None:
    """Tulis sheet 'Duplikat' berisi baris kembar (non-destruktif). index=1 (setelah Ringkasan)."""
    ws = wb.create_sheet(title="Duplikat", index=1)
    headers = ["Grup"] + COLUMNS
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    jumlah_col = headers.index("jumlah") + 1
    for gi, group in enumerate(groups, start=1):
        for row in group:
            ws.append([gi] + [row.get(col, "") for col in COLUMNS])
            cell = ws.cell(row=ws.max_row, column=jumlah_col)
            cell.value = row_jumlah(row)
            cell.number_format = RUPIAH_FMT
        ws.append([None] * len(headers))  # pemisah antar grup

    autofit(ws)


def build_workbook(bills: list[dict]) -> tuple[Workbook, dict]:
    """Bangun workbook dari data. Return (wb, stats)."""
    # Kelompokkan per bulan
    by_month: dict[str, list[dict]] = defaultdict(list)
    for row in bills:
        by_month[month_key(row)].append(row)

    wb = Workbook()
    wb.remove(wb.active)  # buang sheet default kosong

    summary: dict[str, tuple[float, float, int]] = {}
    total_in = total_out = 0.0

    # Saldo kas berjalan, dioper ke tiap sheet bulan sebagai Saldo Awal Bulan.
    # Baris "saldo awal" (kalau nanti ditambahkan ke bills.csv) otomatis jadi titik
    # mulai rantai ini — tidak perlu config terpisah.
    saldo = 0.0
    for bulan in sorted(by_month, key=summary_sort_key):
        rows = by_month[bulan]
        kronologis = bulan != NO_DATE_SHEET
        tin, tout = write_month_sheet(wb, bulan, rows, saldo if kronologis else None)
        if kronologis:
            saldo += tin - tout
        summary[bulan] = (tin, tout, len(rows))
        total_in += tin
        total_out += tout

    write_summary_sheet(wb, summary)

    # Laporan duplikat (non-destruktif): semua baris tetap ada di sheet bulan,
    # sheet "Duplikat" hanya menyorot baris kembar bila ada.
    dup_groups = find_duplicates(bills)
    if dup_groups:
        write_duplicate_sheet(wb, dup_groups)

    stats = {
        "months": len(by_month),
        "transactions": len(bills),
        "total_in": total_in,
        "total_out": total_out,
        "saldo": total_in - total_out,
        "dup_groups": len(dup_groups),
        "dup_rows": sum(len(g) for g in dup_groups),
    }
    return wb, stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Ekspor bills.csv ke Excel per bulan")
    parser.add_argument("--out", help="Path output .xlsx (default: data/backups/bills_export_YYYYMMDD_HHMMSS.xlsx)")
    args = parser.parse_args()

    bills = load_bills()

    if not bills:
        print("⚠️  Tidak ada data di data/bills.csv, tidak ada yang diekspor.")
        return

    wb, stats = build_workbook(bills)

    if args.out:
        out_path = Path(args.out)
    else:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        # Stempel waktu sampai detik → tiap export menghasilkan FILE BARU (tidak menimpa).
        out_path = BACKUP_DIR / f"bills_export_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # Retensi otomatis hanya untuk folder default data/backups/ — jangan sentuh
    # folder custom yang dipilih user via --out.
    deleted = 0 if args.out else cleanup_old_exports(BACKUP_DIR)

    try:
        rel = out_path.relative_to(PROJECT_DIR)
    except ValueError:
        rel = out_path

    print("✅ Export Excel selesai!")
    print(f"📁 File: {rel}")
    print(f"🗓️  Bulan: {stats['months']} sheet  •  📝 Transaksi: {stats['transactions']}")
    print(f"📥 Pemasukan:   {format_rupiah(stats['total_in'])}")
    print(f"📤 Pengeluaran: {format_rupiah(stats['total_out'])}")
    saldo = stats["saldo"]
    print(f"{'💚' if saldo >= 0 else '🔴'} Saldo:       {format_rupiah(saldo)}{'  (MINUS!)' if saldo < 0 else ''}")
    if deleted:
        print(f"🗑️  Retensi: {deleted} export lama (>{RETENSI_HARI} hari) dihapus otomatis.")

    # Laporan duplikat (tanpa menghapus data)
    if stats.get("dup_groups"):
        print(
            f"⚠️  Duplikat: {stats['dup_rows']} baris dalam {stats['dup_groups']} grup "
            f"(lihat sheet 'Duplikat'). Tidak ada data yang dihapus."
        )
    else:
        print("✅ Tidak ada duplikat.")


if __name__ == "__main__":
    main()
